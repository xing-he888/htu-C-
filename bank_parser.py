"""
题库解析器 - 支持多种格式
自动识别并解析：TXT / CSV / Excel / JSON / Markdown
"""

import os
import re
import json
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class BankParser:
    """通用题库解析器，自动识别格式并提取题目-答案对"""

    def __init__(self):
        self.entries = []  # list of {"question": str, "answer": str, "source": str}

    def parse_file(self, filepath: str, source_name: str = None) -> list:
        """
        解析单个文件
        Args:
            filepath: 文件路径
            source_name: 来源名称（用于结果标记）
        Returns:
            list of {"question": str, "answer": str, "source": str}
        """
        if not os.path.exists(filepath):
            logger.warning(f"文件不存在: {filepath}")
            return []

        ext = Path(filepath).suffix.lower()
        source = source_name or os.path.basename(filepath)

        try:
            if ext in [".xlsx", ".xls"]:
                return self._parse_excel(filepath, source)
            elif ext == ".csv":
                return self._parse_csv(filepath, source)
            elif ext == ".json":
                return self._parse_json(filepath, source)
            elif ext in [".md", ".markdown"]:
                return self._parse_markdown(filepath, source)
            elif ext == ".txt":
                return self._parse_txt(filepath, source)
            else:
                # 尝试用文本方式解析
                return self._parse_txt(filepath, source)
        except Exception as e:
            logger.error(f"解析文件失败 {filepath}: {e}")
            return []

    def parse_directory(self, dirpath: str, source_name: str = None) -> list:
        """
        递归解析目录下所有支持的文件
        """
        all_entries = []
        for root, dirs, files in os.walk(dirpath):
            # 跳过 .git 和隐藏目录
            dirs[:] = [d for d in dirs if not d.startswith(".")]
            for f in files:
                if f.startswith("."):
                    continue
                filepath = os.path.join(root, f)
                entries = self.parse_file(filepath, source_name)
                all_entries.extend(entries)

        return all_entries

    # ===== 各格式解析器 =====

    def _parse_txt(self, filepath: str, source: str) -> list:
        """解析文本文件，尝试多种模式"""
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()

        entries = []

        # 模式1: Q: / A: 格式
        qa_pairs = re.findall(
            r"(?:Q|问题|题目)[:：\s]*(.+?)\s*(?:A|答案|解答)[:：\s]*(.+?)(?=(?:Q|问题|题目)[:：]|\Z)",
            content, re.DOTALL | re.IGNORECASE
        )
        for q, a in qa_pairs:
            entries.append({
                "question": q.strip(),
                "answer": a.strip(),
                "source": source
            })

        # 模式2: 编号格式 (1. xxx \n 答案: xxx)
        if not entries:
            numbered = re.findall(
                r"(\d+)[.、)]\s*(.+?)\s*答案[:：]\s*(.+?)(?=\d+[.、)]|\Z)",
                content, re.DOTALL
            )
            for num, q, a in numbered:
                entries.append({
                    "question": q.strip(),
                    "answer": a.strip(),
                    "source": source
                })

        # 模式3: 每行一题，用分隔符
        if not entries:
            lines = content.strip().split("\n")
            current_q = None
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # 检测分隔符
                if "\t" in line or "|||" in line or "---" in line:
                    parts = re.split(r"\t+|\|\|\||---+", line, maxsplit=1)
                    if len(parts) == 2:
                        entries.append({
                            "question": parts[0].strip(),
                            "answer": parts[1].strip(),
                            "source": source
                        })

        return entries

    def _parse_csv(self, filepath: str, source: str) -> list:
        """解析CSV文件"""
        import csv

        entries = []
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            header = next(reader, None)

            # 自动检测题目和答案列
            q_col, a_col = self._detect_qa_columns(header)

            for row in reader:
                if len(row) > max(q_col, a_col):
                    q = row[q_col].strip() if q_col < len(row) else ""
                    a = row[a_col].strip() if a_col < len(row) else ""
                    if q and a:
                        entries.append({
                            "question": q,
                            "answer": a,
                            "source": source
                        })

        return entries

    def _parse_excel(self, filepath: str, source: str) -> list:
        """解析Excel文件"""
        import openpyxl

        entries = []
        wb = openpyxl.load_workbook(filepath, read_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = [str(c) if c else "" for c in rows[0]]
            q_col, a_col = self._detect_qa_columns(header)

            for row in rows[1:]:
                row_vals = [str(c) if c else "" for c in row]
                if len(row_vals) > max(q_col, a_col):
                    q = row_vals[q_col].strip()
                    a = row_vals[a_col].strip()
                    if q and a:
                        entries.append({
                            "question": q,
                            "answer": a,
                            "source": f"{source}/{sheet_name}"
                        })

        wb.close()
        return entries

    def _parse_json(self, filepath: str, source: str) -> list:
        """解析JSON文件"""
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        entries = []

        # 尝试多种JSON结构
        if isinstance(data, list):
            for item in data:
                entry = self._extract_qa_from_dict(item, source)
                if entry:
                    entries.append(entry)
        elif isinstance(data, dict):
            # 可能按章节组织
            for key, value in data.items():
                if isinstance(value, list):
                    for item in value:
                        entry = self._extract_qa_from_dict(item, source)
                        if entry:
                            entries.append(entry)
                elif isinstance(value, dict):
                    entry = self._extract_qa_from_dict(value, source)
                    if entry:
                        entries.append(entry)

        return entries

    def _parse_markdown(self, filepath: str, source: str) -> list:
        """
        解析Markdown文件
        处理C++教材习题的常见格式
        """
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()

        entries = []
        chapter = self._extract_chapter(content, filepath)

        # 策略1: 找 "练习X.X" 或 "习题X.X" 格式
        exercise_pattern = re.compile(
            r"(?:练习|习题|Exercise)\s*(\d+[\.-]\d+)\s*\n+(.+?)(?=(?:练习|习题|Exercise)\s*\d+[\.-]\d+|#{1,4}\s|\Z)",
            re.DOTALL
        )
        for match in exercise_pattern.finditer(content):
            num = match.group(1)
            body = match.group(2)

            q, a = self._split_qa_from_body(body)
            if q:
                entries.append({
                    "question": f"[{num}] {q}",
                    "answer": a or "(答案见原文)",
                    "source": f"{source} > {chapter} > 练习{num}"
                })

        # 策略2: 找编号列表 + 答案
        if not entries:
            entries = self._parse_numbered_qa(content, source, chapter)

        return entries

    # ===== 辅助方法 =====

    def _detect_qa_columns(self, header: list) -> tuple:
        """自动检测题目和答案列索引"""
        q_keywords = ["题目", "问题", "question", "题干", "试题"]
        a_keywords = ["答案", "解答", "answer", "正确", "选项"]

        q_col, a_col = 0, 1  # 默认

        for i, col in enumerate(header):
            col_lower = str(col).lower().strip()
            for kw in q_keywords:
                if kw in col_lower:
                    q_col = i
            for kw in a_keywords:
                if kw in col_lower:
                    a_col = i

        return q_col, a_col

    def _extract_qa_from_dict(self, item: dict, source: str) -> dict:
        """从字典中提取题目-答案"""
        if not isinstance(item, dict):
            return None

        q_keys = ["question", "题目", "问题", "q", "题干"]
        a_keys = ["answer", "答案", "解答", "a", "正确选项"]

        q = None
        a = None

        for k in q_keys:
            if k in item:
                q = str(item[k])
                break

        for k in a_keys:
            if k in item:
                a = str(item[k])
                break

        if q and a:
            return {"question": q, "answer": a, "source": source}
        return None

    def _extract_chapter(self, content: str, filepath: str) -> str:
        """从markdown内容中提取章节名"""
        # 找一级标题
        m = re.search(r"^#\s+(.+)$", content, re.MULTILINE)
        if m:
            return m.group(1)

        # 从文件名推断
        basename = os.path.basename(filepath)
        m = re.search(r"(?:ch|chapter|第)?(\d+)[章章节]", basename, re.IGNORECASE)
        if m:
            return f"第{m.group(1)}章"

        return os.path.splitext(basename)[0]

    def _split_qa_from_body(self, body: str) -> tuple:
        """从题目正文中分离问题和答案"""
        # 找答案标记
        patterns = [
            r"(?:答案|解答|参考答案)[:：]\s*(.+?)$",
            r"(?:Answer|Solution)[:：]\s*(.+?)$",
            r"([A-D][.、)]\s*.+?)$",  # 选择答案
        ]

        for pattern in patterns:
            m = re.search(pattern, body, re.MULTILINE | re.IGNORECASE)
            if m:
                answer = m.group(1).strip()
                question = body[:m.start()].strip()
                return question, answer

        return body.strip(), ""

    def _parse_numbered_qa(self, content: str, source: str, chapter: str) -> list:
        """解析编号形式的QA"""
        entries = []

        # 找 "数字. 题目内容" 后面跟着 "答案: ..." 的模式
        blocks = re.split(r"\n(?=\d+[.、)])", content)

        for block in blocks:
            # 提取编号
            num_m = re.match(r"(\d+)[.、)]", block)
            if not num_m:
                continue
            num = num_m.group(1)

            # 分离问题和答案
            q, a = self._split_qa_from_body(block)
            if q:
                entries.append({
                    "question": f"[{num}] {q}",
                    "answer": a or "(答案见原文)",
                    "source": f"{source} > {chapter} > 题{num}"
                })

        return entries


def test_parser():
    """测试解析器"""
    parser = BankParser()

    # 测试文本解析
    test_content = """1. C++中，下列哪个关键字用于定义类？
答案: class

2. 什么是多态？
答案: 同一操作作用于不同对象，可以有不同的解释，产生不同的执行结果。
"""
    import tempfile
    with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False, encoding="utf-8") as f:
        f.write(test_content)
        tmpfile = f.name

    entries = parser.parse_file(tmpfile, "测试")
    print(f"解析到 {len(entries)} 条题目:")
    for e in entries:
        print(f"  Q: {e['question'][:50]}...")
        print(f"  A: {e['answer'][:50]}...")

    os.unlink(tmpfile)


if __name__ == "__main__":
    test_parser()
